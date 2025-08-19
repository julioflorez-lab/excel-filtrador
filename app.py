import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.title("Filtro ArchivoPlano Tiempo Libre")

# Subir archivo
archivo = st.file_uploader("Sube el archivo Excel", type=["xlsx"])

if archivo is not None:
    # Leer Excel
    df = pd.read_excel(archivo)

    # Convertir fechas (incluyendo la columna Fecha original)
    for col in ["Fecha de pago", "Fecha de factura", "Fecha"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    archivo.seek(0)  # resetear puntero para openpyxl
    wb = load_workbook(archivo, data_only=True)
    ws = wb.active

    colores = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # desde la fila 2 (asumiendo encabezados en fila 1)
        # aquí tomo el color de la columna "Fecha" (ajusta el índice si necesitas otra columna)
        celda = row[df.columns.get_loc("Fecha")]  
        color = celda.fill.start_color.rgb if celda.fill and celda.fill.start_color else None
        colores.append(color)

    df["Color"] = colores

    # Filtro por fecha "Desde" (columna Fecha)
    fecha_desde = None
    if "Fecha" in df.columns:
        fecha_min = df["Fecha"].min()
        fecha_max = df["Fecha"].max()

        fecha_desde = st.date_input(
            "Mostrar registros desde la fecha (columna Fecha):",
            value=None,  # por defecto no aplica filtro
            min_value=fecha_min,
            max_value=fecha_max
        )

    # Filtrar por Pago y ArchivoPlano
    df_filtrado = df[
        (df['Pago'] == True) &
        (df['ArchivoPlano'] == False)
    ]

    colores_unicos = df_filtrado["Color"].dropna().unique().tolist()
    if colores_unicos:
        color_sel = st.selectbox("Filtrar por color de la columna Fecha", ["Todos"] + colores_unicos)
        if color_sel != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Color"] == color_sel]

    # Aplicar filtro de fecha si el usuario selecciona
    if fecha_desde:
        df_filtrado = df_filtrado[
            df_filtrado["Fecha"] >= pd.to_datetime(fecha_desde)
        ]

    # Crear nuevo DataFrame
    df_nuevo = pd.DataFrame()
    df_nuevo['CodigoEntidad'] = df_filtrado['Codigo entidad']
    df_nuevo['Cedula'] = df_filtrado['Identificación']
    df_nuevo['Concepto'] = df_filtrado['Concepto']
    df_nuevo['N° Factura'] = df_filtrado['FacturaSimple'].fillna("").astype(str)
    df_nuevo['Fecha de Pago'] = df_filtrado['Fecha de pago'].dt.strftime("%d/%m/%Y")  # solo Fecha
    df_nuevo['Fecha Factura'] = df_filtrado['Fecha de factura'].dt.strftime("%d/%m/%Y")  # solo Fecha
    df_nuevo['Valor total'] = df_filtrado['Valor total']
    df_nuevo['Valor 0'] = 0
    df_nuevo['Centro de Costo'] = 17395
    df_nuevo['Abreviatura'] = "US"
    df_nuevo['Observacion'] = df_filtrado['Observacion']

    # Mostrar vista previa completa
    st.dataframe(df_nuevo)

    # Mostrar conteo de registros en la interfaz
    total_registros = len(df_nuevo)
    st.subheader(f"📊 Total de registros: {total_registros}")

    # Descargar archivo
    output = BytesIO()
    df_nuevo.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    st.success("Archivo procesado con éxito.")
    st.download_button(
        label="Descargar archivo filtrado",
        data=output,
        file_name="archivo_nuevo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




